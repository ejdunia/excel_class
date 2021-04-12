from openpyxl import load_workbook
import json
import smtplib
import os
from email.message import EmailMessage
# importing all necessary modules to be used

email_add = os.environ.get('MY_EMAIL')
email_pass = os.environ.get('EMAIL_PASS')
# already had set os.environ files 

# loading the excel sheet 
wb = load_workbook('email_test_scores.xlsx')
# print(wb.sheetnames)
sheet = wb.active

# initialise a dict
email_scores = {}
  
# iterating through the excel file to store key value pairs
for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
    emails = row[0]
    grades = row[1]     
    email_scores[emails] = grades


def get_usernames():
    usernames =[u_name[:u_name.find("@")] for u_name in email_scores]
    return usernames

def get_email_list():
    email_list = [emailss for emailss in email_scores]
    return email_list

def get_grades():
    grade = [grade for grade in email_scores.values()]
    return grade

def send_message(e_mesg, subject, email_add, email, message):
    """function to send the mail with  """
    msg = e_mesg
    msg['Subject'] = subject
    msg['From'] = email_add
    msg['To'] = email
    msg.set_content(message)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_add, email_pass)
        smtp.send_message(msg)

get_email_list()
get_grades()
get_usernames()

for (username, email, grde) in zip(get_usernames(), get_email_list(), get_grades()):
    pass_msg = f"Hi {username},\n this message was sent as a result of a refactor test {grde}, therefore..."

    fail_msg = f"Dear {username} I am sorry to inform you that you scored {grde} in your interview test which is less than the pass mark,\nyour journey ends with us here and we wish you all the best"

    def check_draft():
        """a function to set the message to be sent out """
        if grde < 70:
            draft = fail_msg
        else:
            draft = pass_msg
        return draft


    send_message(e_mesg=EmailMessage(), subject="Python Test message refactor test", email_add=email_add, email=email, message=check_draft())
    # send_message("Python test", EmailMessage(), email_add, email, check_draft())

